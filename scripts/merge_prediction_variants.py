from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge two prediction CSV files by sample_type.")
    parser.add_argument("--base-csv", type=Path, required=True)
    parser.add_argument("--override-csv", type=Path, required=True)
    parser.add_argument("--override-types", nargs="+", default=[])
    parser.add_argument("--conditional-override-types", nargs="+", default=[])
    parser.add_argument("--base-strategy-must-be", nargs="+", default=[])
    parser.add_argument("--override-strategy-must-be", nargs="+", default=[])
    parser.add_argument("--template", type=Path, default=ROOT / "决赛答题卡（示例）.xlsx")
    parser.add_argument("--output-csv", type=Path, required=True)
    parser.add_argument("--output-xlsx", type=Path, required=True)
    return parser.parse_args()


def write_answer_workbook(rows: list[dict], template_path: Path, output_path: Path) -> None:
    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    headers = ["题目序号", "后5秒续写", "源文件"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    max_rows = max(ws.max_row, len(rows) + 1)
    for row_idx in range(2, max_rows + 1):
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["题目序号"])
        ws.cell(row=row_idx, column=2, value=row["后5秒续写"])
        ws.cell(row=row_idx, column=3, value=row["源文件"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()
    override_types = set(args.override_types)
    conditional_override_types = set(args.conditional_override_types)
    base_strategy_must_be = set(args.base_strategy_must_be)
    override_strategy_must_be = set(args.override_strategy_must_be)

    base_df = pd.read_csv(args.base_csv)
    override_df = pd.read_csv(args.override_csv)

    key_cols = ["题目序号", "源文件"]
    merged = base_df.merge(
        override_df,
        on=key_cols,
        suffixes=("_base", "_override"),
    )

    rows: list[dict] = []
    output_rows: list[dict] = []
    for _, row in merged.iterrows():
        sample_type = str(row["sample_type_override"])
        use_override = sample_type in override_types

        if not use_override and sample_type in conditional_override_types:
            base_ok = True
            override_ok = True
            if base_strategy_must_be:
                base_ok = str(row.get("generation_strategy_base", "")) in base_strategy_must_be
            if override_strategy_must_be:
                override_ok = str(row.get("generation_strategy_override", "")) in override_strategy_must_be
            use_override = base_ok and override_ok

        suffix = "_override" if use_override else "_base"
        output_row = {}
        for col in merged.columns:
            if col in key_cols:
                output_row[col] = row[col]
            elif col.endswith("_base") or col.endswith("_override"):
                if col.endswith(suffix):
                    output_row[col[: -len(suffix)]] = row[col]
        output_rows.append(output_row)
        rows.append(
            {
                "题目序号": row["题目序号"],
                "后5秒续写": output_row["后5秒续写"],
                "源文件": row["源文件"],
            }
        )

    output_df = pd.DataFrame(output_rows)
    args.output_csv.parent.mkdir(parents=True, exist_ok=True)
    output_df.to_csv(args.output_csv, index=False, encoding="utf-8-sig")
    write_answer_workbook(rows, args.template, args.output_xlsx)

    print(f"merged {len(output_df)} rows")
    print(f"override_types={sorted(override_types)}")
    print(f"conditional_override_types={sorted(conditional_override_types)}")
    if base_strategy_must_be:
        print(f"base_strategy_must_be={sorted(base_strategy_must_be)}")
    if override_strategy_must_be:
        print(f"override_strategy_must_be={sorted(override_strategy_must_be)}")
    print(f"csv: {args.output_csv}")
    print(f"xlsx: {args.output_xlsx}")


if __name__ == "__main__":
    main()
