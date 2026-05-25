from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from gameai2026.retrieval_baseline import (
    RetrievalBaseline,
    collect_records,
    dataframe_to_records,
    records_to_dataframe,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Predict the finals test set and export an answer sheet.")
    parser.add_argument("--train-root", type=Path, default=ROOT / "final_all")
    parser.add_argument("--test-root", type=Path, default=ROOT / "决赛测试100题")
    parser.add_argument("--template", type=Path, default=ROOT / "决赛答题卡（示例）.xlsx")
    parser.add_argument("--train-records", type=Path, default=ROOT / "outputs" / "train_records.csv")
    parser.add_argument("--pred-csv", type=Path, default=ROOT / "outputs" / "test_predictions.csv")
    parser.add_argument("--answer-xlsx", type=Path, default=ROOT / "outputs" / "决赛答题卡-预测版.xlsx")
    parser.add_argument("--train-limit", type=int, default=None)
    parser.add_argument("--test-limit", type=int, default=None)
    parser.add_argument("--top-k", type=int, default=5)
    parser.add_argument("--candidate-pool", type=int, default=60)
    parser.add_argument("--text-weight", type=float, default=0.78)
    parser.add_argument("--numeric-weight", type=float, default=0.22)
    parser.add_argument("--first-action-weight", type=float, default=0.22)
    parser.add_argument("--second-action-weight", type=float, default=0.12)
    parser.add_argument("--enable-action-tail-cls", action="store_true")
    parser.add_argument("--enable-grenade-pair-cls", action="store_true")
    parser.add_argument("--rebuild-train-records", action="store_true")
    return parser.parse_args()


def load_or_build_train_records(args: argparse.Namespace):
    if args.train_records.exists() and not args.rebuild_train_records:
        df = pd.read_csv(args.train_records)
        return dataframe_to_records(df)

    records = collect_records(args.train_root, with_label=True, limit=args.train_limit)
    df = records_to_dataframe(records)
    args.train_records.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(args.train_records, index=False, encoding="utf-8-sig")
    return records


def write_answer_workbook(rows: list[dict], template_path: Path, output_path: Path) -> None:
    if template_path.exists():
        wb = load_workbook(template_path)
        ws = wb[wb.sheetnames[0]]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    headers = ["题目序号", "后5秒续写", "源文件"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    max_rows = max(ws.max_row, len(rows) + 1)
    for row_idx in range(2, max_rows + 1):
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx, value=None)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["题目序号"])
        ws.cell(row=row_idx, column=2, value=row["后5秒续写"])
        ws.cell(row=row_idx, column=3, value=row["源文件"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()

    train_records = load_or_build_train_records(args)
    test_records = collect_records(args.test_root, with_label=False, limit=args.test_limit)

    baseline = RetrievalBaseline(
        text_weight=args.text_weight,
        numeric_weight=args.numeric_weight,
        candidate_pool=args.candidate_pool,
        first_action_weight=args.first_action_weight,
        second_action_weight=args.second_action_weight,
        enable_action_tail_classifier=args.enable_action_tail_cls,
        enable_grenade_pair_classifier=args.enable_grenade_pair_cls,
    ).fit(train_records)

    rows: list[dict] = []
    for idx, record in enumerate(test_records, start=1):
        result = baseline.predict_one(record, top_k=args.top_k)
        best_neighbor = result["neighbors"][0] if result["neighbors"] else {}
        rows.append(
            {
                "题目序号": idx,
                "后5秒续写": result["prediction_text"],
                "源文件": Path(record.sample_path).name,
                "sample_type": record.sample_type,
                "decision_type": record.decision_type,
                "generation_strategy": result.get("generation_strategy", ""),
                "predicted_future_action_1": result.get("predicted_future_action_1", ""),
                "predicted_future_action_2": result.get("predicted_future_action_2", ""),
                "top1_score": float(best_neighbor.get("score", 0.0)),
                "top1_label": best_neighbor.get("label_text", ""),
                "top1_source": best_neighbor.get("sample_path", ""),
                "top1_future_action_text": best_neighbor.get("future_action_text", ""),
            }
        )

    pred_df = pd.DataFrame(rows)
    args.pred_csv.parent.mkdir(parents=True, exist_ok=True)
    pred_df.to_csv(args.pred_csv, index=False, encoding="utf-8-sig")
    write_answer_workbook(rows, args.template, args.answer_xlsx)

    print(f"predicted {len(rows)} samples")
    print(
        "local_structure_flags:"
        f" action_tail={int(args.enable_action_tail_cls)}"
        f" grenade_pair={int(args.enable_grenade_pair_cls)}"
    )
    print(f"csv: {args.pred_csv}")
    print(f"answer: {args.answer_xlsx}")


if __name__ == "__main__":
    main()
