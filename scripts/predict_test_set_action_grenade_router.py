from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from gameai2026.action_grenade_router import (
    FINAL_BASELINE_PARAMS,
    FINAL_TUNED_AG_PARAMS,
    ROUTED_SAMPLE_TYPES,
    filter_records,
    fit_action_grenade_router,
    make_baseline,
)
from gameai2026.retrieval_baseline import collect_records, dataframe_to_records


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Predict the finals test set with the Action/Grenade conditional router.")
    parser.add_argument("--train-root", type=Path, default=ROOT / "final_all")
    parser.add_argument("--test-root", type=Path, default=ROOT / "决赛测试100题")
    parser.add_argument("--template", type=Path, default=ROOT / "决赛答题卡（示例）.xlsx")
    parser.add_argument("--train-records", type=Path, default=ROOT / "outputs" / "train_records_full_with_future.csv")
    parser.add_argument("--pred-csv", type=Path, default=ROOT / "outputs" / "final_predictions_action_grenade_router.csv")
    parser.add_argument("--answer-xlsx", type=Path, default=ROOT / "outputs" / "final_answer_action_grenade_router.xlsx")
    parser.add_argument("--router-oof-csv", type=Path, default=ROOT / "outputs" / "router_action_grenade_oof_full.csv")
    parser.add_argument("--top-k", type=int, default=5)
    parser.add_argument("--router-oof-folds", type=int, default=3)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument("--enable-action-tail-cls", action="store_true")
    parser.add_argument("--enable-grenade-pair-cls", action="store_true")
    parser.add_argument("--rebuild-train-records", action="store_true")
    return parser.parse_args()


def load_or_build_train_records(args: argparse.Namespace):
    if args.train_records.exists() and not args.rebuild_train_records:
        df = pd.read_csv(args.train_records)
        return dataframe_to_records(df)
    return collect_records(args.train_root, with_label=True)


def write_answer_workbook(rows: list[dict], template_path: Path, output_path: Path) -> None:
    if template_path.exists():
        workbook = load_workbook(template_path)
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"

    headers = ["题目序号", "后5秒续写", "源文件"]
    for column_index, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=column_index, value=header)

    max_rows = max(worksheet.max_row, len(rows) + 1)
    for row_index in range(2, max_rows + 1):
        for column_index in range(1, 4):
            worksheet.cell(row=row_index, column=column_index, value=None)

    for row_index, row in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1, value=row["题目序号"])
        worksheet.cell(row=row_index, column=2, value=row["后5秒续写"])
        worksheet.cell(row=row_index, column=3, value=row["源文件"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()

    train_records = load_or_build_train_records(args)
    test_records = collect_records(args.test_root, with_label=False)

    baseline_params = dict(FINAL_BASELINE_PARAMS)
    tuned_params = dict(FINAL_TUNED_AG_PARAMS)
    for params in (baseline_params, tuned_params):
        params["enable_action_tail_classifier"] = args.enable_action_tail_cls
        params["enable_grenade_pair_classifier"] = args.enable_grenade_pair_cls

    stable_baseline = make_baseline(baseline_params).fit(train_records)

    routed_train_records = filter_records(train_records, ROUTED_SAMPLE_TYPES)
    routed_tuned_baseline = make_baseline(tuned_params).fit(routed_train_records)
    router, router_rows = fit_action_grenade_router(
        routed_train_records,
        top_k=args.top_k,
        oof_folds=args.router_oof_folds,
        seed=args.seed,
        route_types=ROUTED_SAMPLE_TYPES,
        base_params=baseline_params,
        tuned_params=tuned_params,
    )

    rows: list[dict] = []
    answer_rows: list[dict] = []

    for index, record in enumerate(test_records, start=1):
        base_result = stable_baseline.predict_one(record, top_k=args.top_k)
        chosen_result = base_result
        chosen_source = "base"
        router_reason = "stable_type"
        router_probability_tuned = 0.0
        router_threshold = 0.5
        tuned_result = None

        if record.sample_type in ROUTED_SAMPLE_TYPES:
            tuned_result = routed_tuned_baseline.predict_one(record, top_k=args.top_k)
            route_info = router.route_prediction(record, base_result, tuned_result)
            chosen_result = route_info["chosen_result"]
            chosen_source = str(route_info["chosen_source"])
            router_reason = str(route_info["router_reason"])
            router_probability_tuned = float(route_info["router_probability_tuned"])
            router_threshold = float(route_info["router_threshold"])

        best_neighbor = chosen_result["neighbors"][0] if chosen_result["neighbors"] else {}
        tuned_prediction = ""
        tuned_strategy = ""
        if tuned_result is not None:
            tuned_prediction = str(tuned_result["prediction_text"])
            tuned_strategy = str(tuned_result.get("generation_strategy", ""))

        row = {
            "题目序号": index,
            "后5秒续写": str(chosen_result["prediction_text"]),
            "源文件": Path(record.sample_path).name,
            "sample_type": record.sample_type,
            "decision_type": record.decision_type,
            "generation_strategy": str(chosen_result.get("generation_strategy", "")),
            "predicted_future_action_1": str(chosen_result.get("predicted_future_action_1", "")),
            "predicted_future_action_2": str(chosen_result.get("predicted_future_action_2", "")),
            "top1_score": float(best_neighbor.get("score", 0.0)),
            "top1_label": best_neighbor.get("label_text", ""),
            "top1_source": best_neighbor.get("sample_path", ""),
            "top1_future_action_text": best_neighbor.get("future_action_text", ""),
            "router_chosen_source": chosen_source,
            "router_reason": router_reason,
            "router_probability_tuned": router_probability_tuned,
            "router_threshold": router_threshold,
            "base_prediction_text": str(base_result["prediction_text"]),
            "base_generation_strategy": str(base_result.get("generation_strategy", "")),
            "tuned_prediction_text": tuned_prediction,
            "tuned_generation_strategy": tuned_strategy,
        }
        rows.append(row)
        answer_rows.append(
            {
                "题目序号": row["题目序号"],
                "后5秒续写": row["后5秒续写"],
                "源文件": row["源文件"],
            }
        )

    pred_df = pd.DataFrame(rows)
    args.pred_csv.parent.mkdir(parents=True, exist_ok=True)
    pred_df.to_csv(args.pred_csv, index=False, encoding="utf-8-sig")

    router_oof_df = pd.DataFrame(
        [
            {
                "sample_type": row.sample_type,
                "file_stem": row.file_stem,
                "base_prediction": row.base_prediction,
                "tuned_prediction": row.tuned_prediction,
                "base_exact": row.base_exact,
                "tuned_exact": row.tuned_exact,
                "preferred_source": row.preferred_source or "",
                "prediction_same": int(row.prediction_same),
                "base_top1_score": row.base_top1_score,
                "tuned_top1_score": row.tuned_top1_score,
            }
            for row in router_rows
        ]
    )
    router_oof_df.to_csv(args.router_oof_csv, index=False, encoding="utf-8-sig")

    write_answer_workbook(answer_rows, args.template, args.answer_xlsx)

    print(f"predicted {len(rows)} samples")
    print(
        "local_structure_flags:"
        f" action_tail={int(args.enable_action_tail_cls)}"
        f" grenade_pair={int(args.enable_grenade_pair_cls)}"
    )
    for sample_type in ROUTED_SAMPLE_TYPES:
        info = router.routers.get(sample_type)
        if info is not None:
            print(
                f"router_{sample_type}: default={info.default_source} "
                f"threshold={info.threshold:.2f} metrics={info.metrics}"
            )
    print(f"csv: {args.pred_csv}")
    print(f"answer: {args.answer_xlsx}")
    print(f"router_oof: {args.router_oof_csv}")


if __name__ == "__main__":
    main()
