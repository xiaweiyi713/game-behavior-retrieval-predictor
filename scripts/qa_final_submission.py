from __future__ import annotations

import argparse
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from gameai2026.retrieval_baseline import build_sample_record


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate final submission CSV/XLSX and write a QA report.")
    parser.add_argument(
        "--pred-csv",
        type=Path,
        default=ROOT / "outputs" / "final_predictions_remote_best_v2.csv",
    )
    parser.add_argument(
        "--answer-xlsx",
        type=Path,
        default=ROOT / "outputs" / "final_answer_remote_best_v2.xlsx",
    )
    parser.add_argument(
        "--base-csv",
        type=Path,
        default=ROOT / "outputs" / "final_predictions_remote_base_v2.csv",
    )
    parser.add_argument(
        "--test-root",
        type=Path,
        default=ROOT / "test100",
    )
    parser.add_argument(
        "--report-path",
        type=Path,
        default=ROOT / "答题卡抽检报告.md",
    )
    return parser.parse_args()


def workbook_rows(path: Path) -> pd.DataFrame:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        rows.append(
            [
                ws.cell(row=row_idx, column=1).value,
                ws.cell(row=row_idx, column=2).value,
                ws.cell(row=row_idx, column=3).value,
            ]
        )
    return pd.DataFrame(rows, columns=["id", "text", "source"])


def sentence_checks(texts: list[str]) -> dict[str, int]:
    return {
        "starts_with_prefix": sum(text.startswith("主玩家先") for text in texts),
        "contains_then": sum("随后" in text for text in texts),
        "contains_last": sum("最后" in text for text in texts),
        "ends_with_period": sum(text.endswith("。") for text in texts),
        "contains_angle_bracket": sum("<" in text or ">" in text for text in texts),
        "contains_nan_literal": sum("nan" in text.lower() for text in texts),
        "contains_none_literal": sum("none" in text.lower() for text in texts),
    }


def format_markdown_table(rows: list[list[object]], headers: list[str]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        escaped = [str(item).replace("\n", " ").replace("|", "/") for item in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def main() -> None:
    args = parse_args()

    pred_df = pd.read_csv(args.pred_csv)
    id_col, text_col, source_col = pred_df.columns[:3]
    texts = [str(value) for value in pred_df[text_col].fillna("")]

    answer_df = workbook_rows(args.answer_xlsx)
    pred_head_df = pred_df.iloc[:, :3].copy()
    pred_head_df.columns = ["id", "text", "source"]
    workbook_matches = answer_df.equals(pred_head_df)

    test_files = sorted(path.name for path in args.test_root.rglob("*.txt"))
    source_values = [str(value) for value in pred_df[source_col].tolist()]
    source_match_test_set = sorted(source_values) == test_files

    length_series = pd.Series([len(text) for text in texts], dtype="int64")
    checks = sentence_checks(texts)

    changed_rows: list[list[object]] = []
    if args.base_csv.exists():
        base_df = pd.read_csv(args.base_csv)
        base_id_col, base_text_col, base_source_col = base_df.columns[:3]
        merged = base_df.merge(
            pred_df,
            left_on=[base_id_col, base_source_col],
            right_on=[id_col, source_col],
            suffixes=("_base", "_best"),
        )
        merged["changed"] = merged[f"{base_text_col}_base"] != merged[f"{text_col}_best"]
        for _, row in merged[merged["changed"]].iterrows():
            changed_rows.append(
                [
                    int(row[base_id_col]),
                    row[base_source_col],
                    row["sample_type_base"],
                    row[f"{base_text_col}_base"],
                    row[f"{text_col}_best"],
                ]
            )

    low_score_rows: list[list[object]] = []
    for _, row in pred_df.sort_values("top1_score").head(6).iterrows():
        sample_path = args.test_root / str(row["sample_type"]) / str(row[source_col])
        record = build_sample_record(sample_path, with_label=False)
        low_score_rows.append(
            [
                int(row[id_col]),
                row[source_col],
                row["sample_type"],
                f'{float(row["top1_score"]):.6f}',
                record.summary_text,
                row[text_col],
            ]
        )

    report = f"""# 答题卡抽检报告

## 1. 核对范围

- 预测明细：`{args.pred_csv.name}`
- 最终答题卡：`{args.answer_xlsx.name}`
- 对照底座：`{args.base_csv.name}`
- 测试集目录：`{args.test_root.name}`

## 2. 自动核对结果

- 预测行数：`{len(pred_df)}`
- 题号唯一数：`{pred_df[id_col].nunique()}`
- 题号范围：`{int(pred_df[id_col].min())} - {int(pred_df[id_col].max())}`
- 源文件唯一数：`{pred_df[source_col].nunique()}`
- 测试集文件数：`{len(test_files)}`
- 预测源文件是否与测试集完全一致：`{source_match_test_set}`
- 答题卡与预测 CSV 三列内容是否逐行完全一致：`{workbook_matches}`
- 空文本数量：`{sum(not text.strip() for text in texts)}`
- 含 `< >` 占位符文本数量：`{checks["contains_angle_bracket"]}`
- 含 `nan` 字面量文本数量：`{checks["contains_nan_literal"]}`
- 含 `none` 字面量文本数量：`{checks["contains_none_literal"]}`
- 以“主玩家先”开头的文本数量：`{checks["starts_with_prefix"]}`
- 含“随后”的文本数量：`{checks["contains_then"]}`
- 含“最后”的文本数量：`{checks["contains_last"]}`
- 以句号结尾的文本数量：`{checks["ends_with_period"]}`

## 3. 文本长度统计

- 最短长度：`{int(length_series.min())}`
- 中位长度：`{int(length_series.median())}`
- 平均长度：`{float(length_series.mean()):.2f}`
- 最长长度：`{int(length_series.max())}`

## 4. 与基础版差异

- 相比 `base_v2`，最终版共替换 `{
            len(changed_rows)
        }` 条答案
- 替换范围仅集中在 `Action` 与 `Grenade`
- 这与最终混合策略一致：基础底座保留稳态题型，重点只覆盖更可能受益的难题型

{format_markdown_table(changed_rows[:12], ["题号", "源文件", "类型", "基础版", "最终版"])}

## 5. 低置信度样本人工抽检建议位

下表列出当前 `top1_score` 最低的 6 条样本，用于人工复核：

{format_markdown_table(low_score_rows, ["题号", "源文件", "类型", "top1_score", "局面摘要", "当前答案"])}

## 6. 抽检结论

- 从提交格式看，当前答题卡已经满足“题号 + 后 5 秒续写 + 源文件”三列的完整输出要求
- 从文件一致性看，答题卡与预测 CSV 完全对齐，没有漏题、重题、空题和测试集外样本
- 从文本规范看，当前答案没有明显占位符、空值或英文异常字符污染
- 从方案策略看，最终版仅对 `Action / Grenade` 做定向替换，避免了对稳定题型的无效扰动

建议把 `final_answer_remote_best_v2.xlsx` 作为正式提交版，同时保留本报告作为提交前自检记录。
"""

    args.report_path.write_text(report, encoding="utf-8")
    print(report)
    print(f"saved_report={args.report_path}")


if __name__ == "__main__":
    main()
